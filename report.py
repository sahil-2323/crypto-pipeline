import urllib.parse
import pandas as pd
from sqlalchemy import create_engine
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# database configuration
RAW_PASSWORD = "PahkjyDXblei9JI8"
SAFE_PASSWORD = urllib.parse.quote_plus(RAW_PASSWORD)
DATABASE_URL = f"postgresql://postgres:{SAFE_PASSWORD}@db.tddwqyyepoadjpofpweg.supabase.co:5432/postgres"

def fetch_data():
    print("Connecting to cloud database to get metrics......")
    try:
        engine = create_engine(DATABASE_URL)
        query = "SELECT timestamp, symbol, name, current_price_usd, market_cap_usd, price_change_percentage_24h FROM crypto_prices ORDER BY timestamp DESC, market_cap_usd DESC LIMIT 10;"
        df = pd.read_sql(query, con= engine)
        return df
    except Exception as e:
        print("Extraction error: {e}")
        return None
    
def build_excel_dashboard(df):
    if df is None or df.empty:
        print("No data available to generate reports :(")
        return
    print("Building automated executive Excel report.....")
    wb = Workbook()
    ws = wb.active
    ws.title = "Crypto Market Summary"

    # show gridlines
    ws.views.sheetView[0].showGridLines= True

    # styling config
    header_fill = PatternFill(start_color="1A252C", end_color="1A252C", fill_type="solid")
    accent_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    green_fill = PatternFill(start_color="D1F7C4", end_color="D1F7C4", fill_type="solid")
    red_fill = PatternFill(start_color="FFD6D6", end_color="FFD6D6", fill_type="solid")

    font_title = Font(name="Calibri", size=16, bold=True, color="1A252C")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Calibri", size=11, bold=False)
    font_bold = Font(name="Calibri", size=11, bold=True)

    thin_border = Border(
        left= Side(style='thin', color='D0D5DD'),
        right= Side(style='thin', color='D0D5DD'),
        top= Side(style='thin', color='D0D5DD'),
        bottom= Side(style='thin', color='D0D5DD')
    )

    # adding title block
    ws['A1'] = "Crypto Automation Pipeline - Execute Report"
    ws['A1'].font = font_title
    ws['A2'] = f"Report generated (UTC): {pd.Timestamp.utcnow().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(name="Calibri", size=10, italic=True, color="667085")

    # adding KPI cards
    # average price cahnge card
    ws['A4'] = "Avg 24h change"
    ws['A4'].font = font_bold
    ws['A4'].fill = accent_fill
    ws['A4'].border = thin_border
    ws['A4'].alignment = Alignment(horizontal="center")

    #formula
    ws['A5']= "=AVERAGE(F8:F17)"
    ws['A5'].font = font_bold
    ws['A5'].border = thin_border
    ws['A5'].alignment = Alignment(horizontal="right")
    ws['A5'].number_format = '0.00"%"'

    # building data table
    headers = ["Timestamp", "Symbol", "Asset Name", "Current Price (USD)", "Market Cap(USD)", "24h Change (%)"]

    # header at row 7
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_num)
        cell.value = header_title
        cell.font = font_header
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # data rows starting from row 8
    for r_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 8):
        for c_idx, value in enumerate (row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = font_data
            cell.border = thin_border

            # formatting numbers based on columns
            if c_idx in [1]: #timestamp
                cell.alignment = Alignment(horizontal="center")
            elif c_idx in [2, 3]: # symbol, name
                cell.alignment = Alignment(horizontal="left")
            elif c_idx == 4: # price
                cell.alignment = Alignment(horizontal="right")
            elif c_idx == 5: # market cap
                cell.number_format = '$#,##0'
                cell.alignment = Alignment(horizontal="right")
            elif c_idx == 6: # 24h change
                cell.number_format = '0.00"%"'
                cell.alignment = Alignment(horizontal="right")
                # automated conditional colour formatting
                if value >= 0:
                    cell.fill = green_fill
                else:
                    cell.fill = red_fill

    # adjusting column widths automatically
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len +3, 12)

    # saving the report
    report_filename = "cryto_executive_report.xlsx"
    wb.save(report_filename)
    print(f"Formatted report saved to ur workplace as '{report_filename}'")

data_df = fetch_data()
build_excel_dashboard(data_df)